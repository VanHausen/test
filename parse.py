from openpyxl import load_workbook

book = load_workbook(filename='/Desktop/Приложение_к_заданию_бек_разработчика.xlsx', use_iterators=True)

sheet = book.get_sheet_names()[0]
worksheet = book.get_sheet_by_name(sheet)

for row in worksheet.iter_rows():
    print(row)

# проверяет последнюю строку
for cell in row:
    print(cell)














